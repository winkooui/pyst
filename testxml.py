from openpyxl import Workbook, load_workbook
from datetime import datetime

def write_log_to_excel(logfile_path, excel_path):
    # 加载已存在的 xlsx 文件或新建一个 Workbook 对象
    wb = load_workbook(excel_path)
    # 获取活动的工作表
    ws = wb.active

    # 读取日志文件
    with open(logfile_path, 'r') as logfile:
        lines = logfile.readlines()

    # 迭代日志内容并写入单元格
    i = 1
    for line in lines:
        # 判断单元格是否为空
        while ws.cell(row=i, column=1).value is not None:
            # 如果单元格不为空，保留原有数据
            i += 1
        ws.cell(row=i, column=1).value = line.strip()
        ws.cell(row=i, column=2).value = datetime.now().strftime("%Y-%m-%d %H:%M")
        i += 1

    # 保存文件
    wb.save(excel_path)


log_path = '../log/log.txt'
exc_path = '../log/log.xlsx'
write_log_to_excel(log_path,exc_path)