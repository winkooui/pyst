# -*- coding: utf-8 -*-
import logging
import openpyxl

def xlsxlog(log_messages):
    # 创建日志对象，创建处理器，给日志对象添加处理器
    logger = logging.getLogger()
    handler = logging.FileHandler('D:\pyselenium\log\log.xlsx', mode='a')
    logger.addHandler(handler)

    # 打开xlsx文件
    xls = openpyxl.load_workbook('D:\pyselenium\log\log.xlsx')
    sheet = xls.active

    # 起始行
    start_row = sheet.max_row + 1
    handlers = logger.handlers

    for handler in handlers:
        if isinstance(handler, logging.FileHandler):
            # 获取日志消息
            log_messages = handler.stream.readlines()

            # 逐行写入到 Excel 文件的第一列中
            for message in log_messages:
                sheet.cell(row=start_row, column=1).value = message.strip()
                start_row += 1
                xls.save("log.xlsx")
xlsxlog('ceshi1')